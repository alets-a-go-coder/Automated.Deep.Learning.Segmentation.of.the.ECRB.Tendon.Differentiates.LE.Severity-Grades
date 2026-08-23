"""
Format segmentation metrics into a manuscript-style Excel table.

Reads segmentation_metrics_all_labels.csv (or runs from in-memory results)
and writes segmentation_metrics_table.xlsx modeled on example_Table.xlsx:
  mean ± SD [minimum, maximum] in each metric cell.

Usage:
  python format_segmentation_metrics_table.py
  python format_segmentation_metrics_table.py --results-csv segmentation_metrics/segmentation_metrics_all_labels.csv
"""

from __future__ import annotations

import argparse
from pathlib import Path

import numpy as np
import pandas as pd

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
except ImportError as exc:
    raise SystemExit(
        "This script requires openpyxl. Install it with:\n"
        "  python -m pip install openpyxl"
    ) from exc

BASE_DIR = Path(__file__).parent

ANATOMY_LABELS = {
    1: "ECRB Tendon",
    2: "CE Muscle",
    3: "Humerus",
    4: "Radius",
    5: "Ulna",
}
DEFAULT_RESULTS_CSV = BASE_DIR / "segmentation_metrics" / "segmentation_metrics_all_labels.csv"
DEFAULT_OUTPUT_XLSX = BASE_DIR / "segmentation_metrics" / "segmentation_metrics_table.xlsx"

TABLE_TITLE = "Table: Automated Segmentation Performance on the Test Set"
FOOTNOTE = (
    "Note.—n = number of subjects in the test set (n = 20). "
    "Metrics are reported as mean ± SD [minimum, maximum] per structure. "
    "Dice and Jaccard indices are unitless (0–1). Hausdorff distance is in image "
    "voxel-spacing units. Volume overlap = 1 − |V_pred − V_gt| / (V_pred + V_gt)."
)

METRIC_COLUMNS = [
    ("dice_score", "Dice", 3),
    ("jaccard_index", "Jaccard (IoU)", 3),
    ("hausdorff_distance", "Hausdorff Distance", 1),
    ("volumetric_overlap", "Volume Overlap", 3),
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Export segmentation metrics to Excel table.")
    parser.add_argument(
        "--results-csv",
        type=Path,
        default=DEFAULT_RESULTS_CSV,
        help=f"Per-subject metrics CSV (default: {DEFAULT_RESULTS_CSV})",
    )
    parser.add_argument(
        "--output-xlsx",
        type=Path,
        default=DEFAULT_OUTPUT_XLSX,
        help=f"Output Excel path (default: {DEFAULT_OUTPUT_XLSX})",
    )
    parser.add_argument(
        "--table-number",
        type=str,
        default="",
        help='Optional table number prefix, e.g. "2" for "Table 2: ..."',
    )
    return parser.parse_args()


def format_mean_sd_range(
    mean: float, sd: float, minimum: float, maximum: float, decimals: int
) -> str:
    if np.isnan(sd):
        sd = 0.0
    return (
        f"{mean:.{decimals}f} ± {sd:.{decimals}f} "
        f"[{minimum:.{decimals}f}, {maximum:.{decimals}f}]"
    )


def build_table_rows(results_df: pd.DataFrame) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for label in sorted(results_df["label"].unique()):
        label_df = results_df[results_df["label"] == label]
        label_int = int(label)
        row: dict[str, object] = {
            "Label": label_int,
            "Structure": ANATOMY_LABELS.get(label_int, f"Label {label_int}"),
            "n": int(label_df["subject_id"].nunique()),
        }
        for col_key, col_title, decimals in METRIC_COLUMNS:
            series = label_df[col_key].astype(float)
            row[col_title] = format_mean_sd_range(
                float(series.mean()),
                float(series.std(ddof=1)) if series.size > 1 else 0.0,
                float(series.min()),
                float(series.max()),
                decimals,
            )
        rows.append(row)
    return rows


def write_excel_table(
    table_rows: list[dict[str, object]],
    output_path: Path,
    title: str,
    footnote: str,
) -> None:
    metric_headers = [title for _, title, _ in METRIC_COLUMNS]
    headers = ["Label", "Structure", "n", *metric_headers]

    wb = Workbook()
    ws = wb.active
    ws.title = "Segmentation Metrics"

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(bold=True, size=12)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Header row
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Data rows
    for row_idx, row_data in enumerate(table_rows, start=3):
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data[header])
            align = "left" if header == "Structure" else "center"
            cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)

    # Footnote
    footnote_row = 3 + len(table_rows)
    ws.merge_cells(
        start_row=footnote_row,
        start_column=1,
        end_row=footnote_row,
        end_column=len(headers),
    )
    note_cell = ws.cell(row=footnote_row, column=1, value=footnote)
    note_cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Column widths
    widths = [8, 22, 6, 28, 28, 28, 28]
    for col_idx, width in enumerate(widths[: len(headers)], start=1):
        col_letter = ws.cell(row=2, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = width

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main() -> None:
    args = parse_args()
    if not args.results_csv.exists():
        raise FileNotFoundError(f"Results CSV not found: {args.results_csv}")

    results_df = pd.read_csv(args.results_csv)
    table_rows = build_table_rows(results_df)

    title = TABLE_TITLE
    if args.table_number.strip():
        title = f"Table {args.table_number.strip()}: Automated Segmentation Performance on the Test Set"

    write_excel_table(table_rows, args.output_xlsx, title, FOOTNOTE)

    print(f"Excel table: {args.output_xlsx}")
    print()
    print(pd.DataFrame(table_rows).to_string(index=False))


if __name__ == "__main__":
    main()
